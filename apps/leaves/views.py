from django.http import HttpResponse, JsonResponse
from django.views import View
from apps.leaves.models import LeaveRequest
from apps.employees.models import Employee
import openpyxl
from openpyxl.utils import get_column_letter

class ExportLeavesView(View):
    def get(self, request, *args, **kwargs):
        auth_header = request.META.get('HTTP_AUTHORIZATION')
        token = None
        if auth_header and auth_header.startswith('JWT '):
            token = auth_header.split(' ')[1]
        else:
            token = request.GET.get('token')
            
        if not token:
            return HttpResponse("Accès refusé. Jeton d'authentification manquant.", status=403)
            
        try:
            import jwt
            from django.conf import settings
            from apps.employees.models import Employee
            
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            username = payload.get('username')
            if not username:
                return HttpResponse("Jeton invalide.", status=403)
                
            user = Employee.objects.get(username=username)
            if user.role not in [Employee.Role.RH, Employee.Role.ADMIN, Employee.Role.DIRECTEUR_GENERAL]:
                return HttpResponse("Accès refusé. Rôle insuffisant.", status=403)
                
        except jwt.ExpiredSignatureError:
            return HttpResponse("Jeton expiré.", status=403)
        except jwt.InvalidTokenError:
            return HttpResponse("Jeton invalide.", status=403)
        except Employee.DoesNotExist:
            return HttpResponse("Utilisateur introuvable.", status=403)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rapport des Congés"
        
        headers = ["Matricule", "Nom", "Prénom", "Département", "Solde Actuel", "Type de Congé", "Date Début", "Date Fin", "Jours Décomptés", "Statut"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = header
            
        requests = LeaveRequest.objects.all().order_by('-created_at')
        
        row_num = 2
        for req in requests:
            emp = req.employee
            dept_name = emp.department.nom if emp.department else ""
            
            ws.cell(row=row_num, column=1, value=emp.matricule)
            ws.cell(row=row_num, column=2, value=emp.last_name)
            ws.cell(row=row_num, column=3, value=emp.first_name)
            ws.cell(row=row_num, column=4, value=dept_name)
            ws.cell(row=row_num, column=5, value=float(emp.solde_conges))
            ws.cell(row=row_num, column=6, value=req.leave_type.libelle)
            ws.cell(row=row_num, column=7, value=req.date_debut.strftime("%Y-%m-%d"))
            ws.cell(row=row_num, column=8, value=req.date_fin.strftime("%Y-%m-%d"))
            ws.cell(row=row_num, column=9, value=float(req.jours_decomptes))
            ws.cell(row=row_num, column=10, value=req.get_statut_display())
            
            row_num += 1
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=rapport_conges.xlsx'
        wb.save(response)
        
        return response

import jwt
from django.conf import settings
from django.http import Http404

class SecureDocumentView(View):
    def get(self, request, *args, **kwargs):
        token = request.GET.get('token')
        if not token:
            return HttpResponse("Jeton manquant.", status=403)
            
        try:
            payload = jwt.decode(token, settings.SECRET_KEY, algorithms=['HS256'])
            leave_request_id = payload['leave_request_id']
            user_id = payload.get('user_id')
            
            leave_request = LeaveRequest.objects.get(id=leave_request_id)
            if not leave_request.piece_jointe:
                raise Http404("Document introuvable")
                
            # Serve the file securely
            import mimetypes
            import os
            
            file_path = leave_request.piece_jointe.path
            if not os.path.exists(file_path):
                raise Http404("Fichier introuvable sur le disque")
                
            content_type, _ = mimetypes.guess_type(file_path)
            if not content_type:
                content_type = 'application/octet-stream'
                
            # Log the access
            from apps.core.models import AuditLog
            AuditLog.objects.create(
                leave_request=leave_request,
                action=AuditLog.Action.ACCESS_DOCUMENT,
                user_id=user_id,
                details=f"Accès sécurisé à la pièce jointe: {os.path.basename(file_path)}"
            )
                
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type=content_type)
                # Inline disposition allows viewing in browser (e.g. PDF)
                response['Content-Disposition'] = f'inline; filename="{os.path.basename(file_path)}"'
                return response
                
        except jwt.ExpiredSignatureError:
            return HttpResponse("Le lien a expiré.", status=403)
        except jwt.InvalidTokenError:
            return HttpResponse("Lien invalide.", status=403)
        except LeaveRequest.DoesNotExist:
            raise Http404("Demande introuvable")
        except Exception as e:
            return HttpResponse(str(e), status=500)
